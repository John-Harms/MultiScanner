import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk
from openpyxl import load_workbook
import base64
import requests
import re
import threading


class PictureSelectorApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Picture Selector")

        # Set the dark theme
        self.style = ttk.Style()
        self.style.theme_use('clam')

        # Set the background color of the root window
        self.master.configure(background='#333333')

        # Configure the styles for ttk widgets
        self.style.configure('TFrame', background='#333333')
        self.style.configure('TLabelFrame', background='#333333', foreground='#FFFFFF')
        self.style.configure('TLabel', background='#333333', foreground='#FFFFFF')
        self.style.configure('TButton', background='#555555', foreground='#FFFFFF')
        self.style.configure('TEntry', fieldbackground='#555555', foreground='#FFFFFF')
        self.style.configure('Vertical.TScrollbar', background='#333333', troughcolor='#555555')
        self.style.configure('Horizontal.TScrollbar', background='#333333', troughcolor='#555555')

        # List to hold selected picture file paths
        self.selected_pictures = []

        # Main frame
        self.main_frame = ttk.Frame(self.master)
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # Create frames for different sections
        self.create_widgets()

    def create_widgets(self):
        # Configure grid weights
        self.master.columnconfigure(0, weight=1)
        self.master.rowconfigure(0, weight=1)
        self.main_frame.columnconfigure(0, weight=1)
        self.main_frame.columnconfigure(1, weight=1)

        # Picture selection and display frame
        self.picture_frame = ttk.LabelFrame(self.main_frame, text="Selected Pictures")
        self.picture_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

        # Entries frame
        self.entry_frame = ttk.LabelFrame(self.main_frame, text="Criteria Entries")
        self.entry_frame.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")

        # Action buttons frame
        self.action_frame = ttk.Frame(self.main_frame)
        self.action_frame.grid(row=1, column=0, columnspan=2, pady=10)

        # Configure frames to expand
        self.picture_frame.columnconfigure(0, weight=1)
        self.picture_frame.rowconfigure(0, weight=1)
        self.entry_frame.columnconfigure(0, weight=1)
        self.entry_frame.rowconfigure(0, weight=1)

        # Picture canvas and scrollbar
        self.picture_canvas = tk.Canvas(self.picture_frame)
        self.picture_scrollbar = ttk.Scrollbar(
            self.picture_frame, orient="vertical", command=self.picture_canvas.yview
        )
        self.picture_canvas.configure(yscrollcommand=self.picture_scrollbar.set)
        self.picture_canvas.grid(row=0, column=0, sticky="nsew")
        self.picture_scrollbar.grid(row=0, column=1, sticky="ns")

        # Set canvas background to dark
        self.picture_canvas.configure(bg='#333333', highlightbackground='#333333')

        self.picture_scrollable_frame = ttk.Frame(self.picture_canvas)
        self.picture_canvas.create_window((0, 0), window=self.picture_scrollable_frame, anchor="nw")
        self.picture_scrollable_frame.bind(
            "<Configure>",
            lambda e: self.picture_canvas.configure(scrollregion=self.picture_canvas.bbox("all"))
        )

        # Entry canvas and scrollbar
        self.entry_canvas = tk.Canvas(self.entry_frame)
        self.entry_scrollbar = ttk.Scrollbar(
            self.entry_frame, orient="vertical", command=self.entry_canvas.yview
        )
        self.entry_canvas.configure(yscrollcommand=self.entry_scrollbar.set)
        self.entry_canvas.grid(row=0, column=0, sticky="nsew")
        self.entry_scrollbar.grid(row=0, column=1, sticky="ns")

        # Set canvas background to dark
        self.entry_canvas.configure(bg='#333333', highlightbackground='#333333')

        self.entry_scrollable_frame = ttk.Frame(self.entry_canvas)
        self.entry_canvas.create_window((0, 0), window=self.entry_scrollable_frame, anchor="nw")
        self.entry_scrollable_frame.bind(
            "<Configure>",
            lambda e: self.entry_canvas.configure(scrollregion=self.entry_canvas.bbox("all"))
        )

        # Entries list
        self.entries = []

        # Add entry button
        self.add_entry_button = ttk.Button(
            self.entry_frame, text="+ Add Entry", command=self.add_entry
        )
        self.add_entry_button.grid(row=1, column=0, pady=5, sticky="ew")

        # Action buttons
        self.select_button = ttk.Button(
            self.action_frame, text="Select Pictures", command=self.select_pictures
        )
        self.select_button.pack(side=tk.LEFT, padx=5)

        self.check_button = ttk.Button(
            self.action_frame, text="Check Entries", command=self.check_entries
        )
        self.check_button.pack(side=tk.LEFT, padx=5)

        self.submit_button = ttk.Button(
            self.action_frame, text="Submit", command=self.submit
        )
        self.submit_button.pack(side=tk.LEFT, padx=5)

        # Custom progress bar canvas
        self.progress_canvas = tk.Canvas(self.action_frame, width=200, height=20, bg='grey')
        self.progress_canvas.pack(side=tk.LEFT, padx=5)

    def select_pictures(self):
        # Open file dialog to select pictures
        file_paths = filedialog.askopenfilenames(
            title="Select Pictures",
            filetypes=[("Image Files", "*.png;*.jpg;*.jpeg;*.gif;*.bmp")],
        )
        if file_paths:
            for file_path in file_paths:
                if len(self.selected_pictures) < 5:
                    if file_path not in self.selected_pictures:
                        self.selected_pictures.append(file_path)
                else:
                    messagebox.showwarning(
                        "Limit Reached", "You can select no more than 5 pictures."
                    )
                    break
            self.display_selected_pictures()

    def display_selected_pictures(self):
        # Clear current displayed pictures
        for widget in self.picture_scrollable_frame.winfo_children():
            widget.destroy()

        for file_path in self.selected_pictures:
            frame = ttk.Frame(self.picture_scrollable_frame)
            frame.pack(pady=5, anchor="w")

            # Load image
            try:
                image = Image.open(file_path)
                image.thumbnail((100, 100))  # Resize image to thumbnail size
                photo = ImageTk.PhotoImage(image)

                label = ttk.Label(frame, image=photo)
                label.image = photo  # Keep a reference
                label.pack(side="left")
            except Exception as e:
                print(f"Error loading image {file_path}: {e}")
                continue

            # Deselect button
            deselect_button = ttk.Button(
                frame,
                text="Deselect",
                command=lambda fp=file_path: self.deselect_picture(fp),
            )
            deselect_button.pack(side="left", padx=5)

    def deselect_picture(self, file_path):
        if file_path in self.selected_pictures:
            self.selected_pictures.remove(file_path)
            self.display_selected_pictures()

    def add_entry(self):
        # Create a frame to hold the entry and delete button
        entry_frame = ttk.Frame(self.entry_scrollable_frame)
        entry_frame.pack(pady=5, anchor="w")

        entry = ttk.Entry(entry_frame, width=30)
        entry.pack(side="left")

        def delete_entry():
            entry_frame.destroy()
            # Remove the entry from the entries list
            self.entries.remove(entry)

        delete_button = ttk.Button(entry_frame, text="X", command=delete_entry)
        delete_button.pack(side="left", padx=5)

        self.entries.append(entry)

    def check_entries(self):
        entries_list = [entry.get() for entry in self.entries if entry.get()]
        if entries_list:
            messagebox.showinfo("Entries", "\n".join(entries_list))
        else:
            messagebox.showwarning("No Entries", "No criteria entries found.")
        return entries_list

    def submit(self, event=None):
        # Disable the submit button to prevent multiple submissions
        self.submit_button.config(state="disabled")
        # Reset progress tracking variables
        self.total_images = len(self.selected_pictures)
        self.successful_processings = 0  # Reset the success counter
        self.processed_images = 0  # Reset the processed images counter
        # Clear the progress canvas
        self.progress_canvas.delete('all')
        threading.Thread(target=self.process_submission).start()

    def update_progress(self, success):
        # Calculate the width of each segment
        bar_width = 200  # Width of the progress canvas
        total_images = self.total_images
        segment_width = bar_width / total_images

        # Calculate x position of the segment
        x1 = (self.processed_images - 1) * segment_width
        x2 = self.processed_images * segment_width

        color = 'green' if success else 'red'

        # Draw the rectangle representing the processing result
        self.progress_canvas.create_rectangle(x1, 0, x2, 20, fill=color, outline='')

    def process_submission(self):
        entries_list = [entry.get() for entry in self.entries if entry.get()]
        entries_parsed = ", ".join(entries_list)
        if entries_list:
            if not self.selected_pictures:
                # Show error message in main thread
                self.master.after(0, lambda: messagebox.showerror("Error", "No pictures selected."))
                # Re-enable the submit button
                self.master.after(0, lambda: self.submit_button.config(state="normal"))
                return
            elif 0 < len(self.selected_pictures) <= 5:
                for picture in self.selected_pictures:
                    encoded_image = self.encode_image_to_base64(picture)
                    success = False  # Assume failure initially
                    if encoded_image:
                        print(
                            f"Size in bytes of the image is: {len(encoded_image) / 1024:.2f} KB"
                        )
                        print("API Call occurring, please wait.")
                        message_content = self.apicall(encoded_image, entries_parsed)
                        if message_content:
                            returners, remaining_matches = parse_string(
                                message_content, entries_list
                            )

                            while returners:
                                # Pass the variables to update the Excel sheet
                                update_excel(returners)

                                if remaining_matches:
                                    # Parse the remaining matches
                                    returners, remaining_matches = parse_string(
                                        "[" + "][".join(remaining_matches) + "]",
                                        entries_list,
                                    )
                                else:
                                    break

                            if not returners and not remaining_matches:
                                print(
                                    f"Parsing failed for file {picture}, could not find given criteria."
                                )
                                success = False
                            else:
                                success = True
                        else:
                            print(f"Failed to process image {picture}")
                    else:
                        print(f"Failed to encode image {picture}")
                    # Update counts
                    if success:
                        self.successful_processings += 1
                    self.processed_images += 1
                    # Update the progress bar
                    self.master.after(0, self.update_progress, success)
                # Re-enable the submit button after processing all images
                self.master.after(0, lambda: self.submit_button.config(state="normal"))
            else:
                # Show warning message in main thread
                self.master.after(
                    0, lambda: messagebox.showwarning("Limit Exceeded", "You can submit up to 5 pictures.")
                )
                # Re-enable the submit button
                self.master.after(0, lambda: self.submit_button.config(state="normal"))
        else:
            # Show warning message in main thread
            self.master.after(0, lambda: messagebox.showwarning("No Criteria", "Please input criteria."))
            # Re-enable the submit button
            self.master.after(0, lambda: self.submit_button.config(state="normal"))

    def encode_image_to_base64(self, image_path):
        print("Encoding Image")
        try:
            with open(image_path, "rb") as image_file:
                return base64.b64encode(image_file.read()).decode("utf-8")
        except Exception as e:
            print(f"Error encoding image {image_path}: {e}")
            return None

    def apicall(self, encoded_image, entries_parsed):
        rule_instructions = (
                f"1. Analyze the picture to extract only the data matching the specified criteria: {entries_parsed}.\n"
                "2. For each item in the picture (e.g., a person or object), record the data in the format:\n"
                "   [Criterion1][Criterion2][...][CriterionN] (including the brackets).\n"
                "3. Ensure the data is accurate. If a criterion cannot be confidently identified, use [NA] for that criterion.\n"
                "   Do not guess.\n"
                "4. If only partial criteria are available for an item, include the identified criteria and fill in [NA] for any missing ones.\n"
                "5. Separate each item's data with a new line to clearly distinguish between different items.\n"
                "\n"
                "Example:\n"
                "- Criteria: (phone number, name, age).\n"
                "- Context: A picture containing data entires about people. Person 2's data is partial.\n"
                "\n"
                "Response Example:\n"
                "[\n"
                "    \"[Person 1's phone number][Person 1's name][Person 1's age]\",\n"
                "    \"[NA][Person 2's name][NA]\"\n"
                "]"

        )

        #           ********  PLACE YOUR API KEY IN THE AUTHORIZATION LINE BELOW THIS TEXT  *******
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {'PLACE YOUR API KEY HERE'}",
        }

        question = ""
        payload = {
            "model": "gpt-4o",
            "messages": [
                {"role": "system", "content": f"{rule_instructions}."},
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": f"{question}"
                        },
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:image/jpeg;base64,{encoded_image}",
                                "detail": "auto"
                            }
                        }
                    ]
                }
            ],
            "max_tokens": 300,
            "temperature": 0
        }

        try:
            response = requests.post(
                "https://api.openai.com/v1/chat/completions",
                headers=headers,
                json=payload,
            )
            response.raise_for_status()
            response_dict = response.json()
            message_content = response_dict["choices"][0]["message"]["content"]

            # Printing the message
            print("AI Response Message:")
            print(message_content)

            usage_info = response_dict.get("usage", {})
            if usage_info:
                print("Usage Information:")
                print(f"Prompt Tokens: {usage_info.get('prompt_tokens', 0)}")
                print(f"Completion Tokens: {usage_info.get('completion_tokens', 0)}")
                print(f"Total Tokens: {usage_info.get('total_tokens', 0)}")

            return message_content
        except requests.exceptions.RequestException as e:
            print(f"API request failed: {e}")
            return ""
        except KeyError:
            print("Unexpected response format.")
            return ""


def parse_string(s, entries_list):
    # Use regular expression to find all content within brackets
    matches = re.findall(r"\[(.*?)\]", s)
    returners = []
    if matches:
        for i in range(len(entries_list)):
            if i < len(matches):
                returners.append(matches[i])
            else:
                returners.append("NA")
        if len(matches) > len(entries_list):
            remaining_matches = matches[len(entries_list):]
            return returners, remaining_matches
        else:
            return returners, None
    else:
        # Handle cases where the expected pattern is not found
        return returners, None


#           *********** PLACE YOUR EXCEL FILE LOCATION DOWN HERE WITHIN THE QUOTES *********

def update_excel(values, sheetname="Sheet1"):
    # Load the workbook
    filename = r'PLACE YOUR EXCEL FILE LOCATION HERE'
    wb = load_workbook(filename)

    # Get the sheet by name
    ws = wb[sheetname]

    # Find the topmost empty cell in column A
    row = 1
    while ws.cell(row=row, column=1).value is not None:
        row += 1

    for col, value in enumerate(values, start=1):
        ws.cell(row=row, column=col).value = value

    # Save the workbook
    wb.save(filename)


if __name__ == "__main__":
    root = tk.Tk()
    app = PictureSelectorApp(root)
    root.mainloop()
